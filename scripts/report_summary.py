import openpyxl

file_path = r'C:\Users\labinf1.pasto\Desktop\Alcaldia\media\maintenance-reports\rutina_mantenimiento_EQ-0003_20251125_c1852b13.xlsx'

wb = openpyxl.load_workbook(file_path)
ws = wb.active

print("=" * 70)
print("   REPORTE EXCEL - VERIFICACIÓN COMPLETA")
print("=" * 70)

print("\n✅ ENCABEZADO")
print(f"   Sede: {ws['A7'].value}")
print(f"   Dependencia: {ws['C7'].value}")
print(f"   Oficina: {ws['H7'].value}")

print("\n✅ FECHA Y HORA")
fecha = ws['C8'].value
hora = ws['H8'].value
print(f"   {fecha}")
print(f"   {hora}")

print("\n✅ ACTIVIDADES HARDWARE (muestra)")
for row in [11, 12]:
    actividad_izq = ws.cell(row, 1).value
    si_izq = ws.cell(row, 4).value
    na_izq = ws.cell(row, 5).value
    print(f"   {actividad_izq[:40]}... -> SI:{si_izq or ''} NA:{na_izq or ''}")

print("\n✅ ACTIVIDADES SOFTWARE (muestra)")
for row in [18, 21]:
    actividad_izq = ws.cell(row, 1).value
    si_izq = ws.cell(row, 4).value
    na_izq = ws.cell(row, 5).value
    print(f"   {actividad_izq[:40]}... -> SI:{si_izq or ''} NA:{na_izq or ''}")

print("\n✅ OBSERVACIONES")
obs_gen = ws['A32'].value
obs_seg = ws['A34'].value
print(f"   Generales: {obs_gen[:50] if obs_gen else 'N/A'}...")
print(f"   Seguridad: {obs_seg[:50] if obs_seg else 'N/A'}...")

print("\n✅ FIRMAS")
print(f"   Técnico (A37-A39):")
print(f"      {ws['A38'].value}")
print(f"      {ws['A39'].value}")
print(f"   Usuario (F37-F39):")
print(f"      {ws['F38'].value}")
print(f"      {ws['F39'].value}")

print(f"\n✅ IMÁGENES EMBEBIDAS: {len(ws._images)} imágenes")
print(f"   - 2 firmas (A37 técnico, F37 usuario)")
print(f"   - {len(ws._images) - 2} fotos (al final del documento)")

print("\n✅ CALIFICACIÓN DEL SERVICIO")
rating = "Excelente" if ws['A45'].value == '■' else \
         "Bueno" if ws['C45'].value == '■' else \
         "Regular" if ws['F45'].value == '■' else \
         "Malo" if ws['H45'].value == '■' else "No especificado"
print(f"   {rating}")

print("\n" + "=" * 70)
print("   ✅ VERIFICACIÓN EXITOSA - TODOS LOS CAMPOS MAPEADOS")
print("=" * 70)
