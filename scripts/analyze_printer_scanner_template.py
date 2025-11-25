import openpyxl
import sys

template_path = r'C:\Users\labinf1.pasto\Desktop\Alcaldia\plantillas\Mantenimiento Escaneres e impresoras\rutina_mantenimiento_preventivo_impresoras_y_escaner_v2.xlsx'
mapeo_path = r'C:\Users\labinf1.pasto\Desktop\Alcaldia\plantillas\Mantenimiento Escaneres e impresoras\Mapeo_rutina_mantenimiento_preventivo_impresoras_y_escaner_v2.xlsx'

print("=" * 80)
print("ANÁLISIS PLANTILLA: IMPRESORAS Y ESCÁNERES")
print("=" * 80)

wb_template = openpyxl.load_workbook(template_path)
ws_template = wb_template.active

print(f"\nHoja: {ws_template.title}")
print(f"Dimensiones: {ws_template.dimensions}")

wb_mapeo = openpyxl.load_workbook(mapeo_path)
ws_mapeo = wb_mapeo.active

# Comparar para encontrar diferencias (zonas marcadas)
differences = []

for row in range(1, min(80, ws_template.max_row + 1)):
    for col in range(1, 11):  # A-J
        cell_template = ws_template.cell(row, col)
        cell_mapeo = ws_mapeo.cell(row, col)
        
        val_template = cell_template.value
        val_mapeo = cell_mapeo.value
        
        if val_template != val_mapeo:
            cell_coord = ws_template.cell(row, col).coordinate
            differences.append({
                'cell': cell_coord,
                'row': row,
                'col': col,
                'original': val_template,
                'mapeo': val_mapeo
            })

print(f"\n{'='*80}")
print(f"TOTAL DE DIFERENCIAS ENCONTRADAS: {len(differences)}")
print(f"{'='*80}")

# Agrupar por secciones
header_cells = [d for d in differences if d['row'] <= 10]
activities_cells = [d for d in differences if 11 <= d['row'] <= 35]
obs_cells = [d for d in differences if 36 <= d['row'] <= 40]
firma_cells = [d for d in differences if 41 <= d['row'] <= 50]
rating_cells = [d for d in differences if 51 <= d['row'] <= 55]
final_cells = [d for d in differences if d['row'] >= 56]

print(f"\n📋 ENCABEZADO (filas 1-10): {len(header_cells)} celdas a rellenar")
for d in header_cells:
    print(f"   {d['cell']}: '{d['original']}' -> MARCADO: '{d['mapeo']}'")

print(f"\n⚙️ ACTIVIDADES (filas 11-35): {len(activities_cells)} celdas")
if activities_cells:
    print("   Muestra de primeras 10:")
    for d in activities_cells[:10]:
        desc = str(d['original'])[:60] if d['original'] else ''
        print(f"   {d['cell']}: {desc}...")

print(f"\n📝 OBSERVACIONES (filas 36-40): {len(obs_cells)} celdas")
for d in obs_cells:
    print(f"   {d['cell']}: marcador='{d['mapeo']}'")

print(f"\n✍️ FIRMAS (filas 41-50): {len(firma_cells)} celdas")
for d in firma_cells:
    print(f"   {d['cell']}: '{d['original']}' -> '{d['mapeo']}'")

print(f"\n⭐ CALIFICACIÓN (filas 51-55): {len(rating_cells)} celdas")
for d in rating_cells:
    print(f"   {d['cell']}: marcador='{d['mapeo']}'")

print(f"\n📸 ZONA FINAL/FOTOS (filas 56+): {len(final_cells)} celdas")
for d in final_cells[:5]:
    print(f"   {d['cell']}: marcador='{d['mapeo']}'")

# Análisis detallado de estructura de actividades
print(f"\n{'='*80}")
print("ESTRUCTURA DE ACTIVIDADES")
print(f"{'='*80}")

activity_rows = {}
for row in range(11, 36):
    a_val = ws_template.cell(row, 1).value
    d_val = ws_template.cell(row, 4).value
    e_val = ws_template.cell(row, 5).value
    f_val = ws_template.cell(row, 6).value
    i_val = ws_template.cell(row, 9).value
    j_val = ws_template.cell(row, 10).value
    
    if a_val or f_val:
        activity_rows[row] = {
            'left': a_val,
            'right': f_val,
            'has_checkboxes': bool(d_val is None and e_val is None)
        }

print(f"\nFilas de actividades encontradas: {len(activity_rows)}")
for row, data in list(activity_rows.items())[:15]:
    left = str(data['left'])[:40] if data['left'] else ''
    right = str(data['right'])[:40] if data['right'] else ''
    print(f"   Fila {row}: IZQ='{left}' | DER='{right}'")

# Verificar zona de firmas específicamente
print(f"\n{'='*80}")
print("VERIFICACIÓN ZONA DE FIRMAS")
print(f"{'='*80}")
for row in range(41, 51):
    for col_letter in ['A', 'F']:
        cell_ref = f'{col_letter}{row}'
        val_t = ws_template[cell_ref].value
        val_m = ws_mapeo[cell_ref].value
        if val_t or val_m:
            match = '✓' if val_t == val_m else '■ MARCADO'
            print(f"{cell_ref}: Template='{val_t}' | Mapeo='{val_m}' [{match}]")

print(f"\n{'='*80}")
print("✅ ANÁLISIS COMPLETADO")
print(f"{'='*80}")
