from rest_framework.decorators import api_view, parser_classes, permission_classes
from django.views.decorators.csrf import csrf_exempt
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from django.http import HttpResponse
from .models import Template
from .models import ReportTemplate
from .models import Report
import json
from datetime import datetime

try:
    from .services.html_pdf_generator import HTMLPDFGenerator
except Exception:
    HTMLPDFGenerator = None

# Fallback simple generator using ReportLab (pure Python). This avoids
# native deps on environments like Windows. reportlab must be installed
# in the virtualenv for this to work.
try:
    from .services.reportlab_pdf_generator import ReportLabPDFGenerator
except Exception:
    ReportLabPDFGenerator = None

import json
from datetime import datetime


# Helper: normalization and auto-mapping used across endpoints
def _norm_key(s):
    try:
        return ''.join([c for c in str(s).lower() if c.isalnum()])
    except Exception:
        return str(s).lower()


def _auto_map_value_for_dict(map_to, tpl_key, data_dict, resolve_path_fn):
    # Try explicit path first
    if map_to:
        v = resolve_path_fn(data_dict, map_to) if isinstance(data_dict, (dict, list)) else None
        if v is not None:
            return v
        if isinstance(data_dict, dict) and map_to in data_dict:
            return data_dict.get(map_to)

    # Heuristic fallback: try to match tpl_key against available keys
    if not isinstance(data_dict, dict):
        return None

    tpl_norm = _norm_key(tpl_key)
    # 1) exact match ignoring case/characters
    for k in data_dict.keys():
        if _norm_key(k) == tpl_norm:
            return data_dict.get(k)

    # 2) substring match (tpl in key or key in tpl)
    for k in data_dict.keys():
        kn = _norm_key(k)
        if tpl_norm in kn or kn in tpl_norm:
            return data_dict.get(k)

    # 3) token overlap
    for k in data_dict.keys():
        k_tokens = set([t for t in _norm_key(k).split() if t])
        t_tokens = set([t for t in tpl_norm.split() if t])
        if k_tokens & t_tokens:
            return data_dict.get(k)

    return None


def _find_matching_key(map_to, tpl_key, data_dict, resolve_path_fn):
    # Returns the matched key name from data_dict if any, else None
    if map_to:
        # if map_to exists exactly in dict
        if isinstance(data_dict, dict) and map_to in data_dict:
            return map_to
        # if path resolves to a value, we can't determine the exact key easily
        # so continue to heuristics

    if not isinstance(data_dict, dict):
        return None

    tpl_norm = _norm_key(tpl_key)
    # 1) exact match ignoring case/characters
    for k in data_dict.keys():
        if _norm_key(k) == tpl_norm:
            return k

    # 2) substring match
    for k in data_dict.keys():
        kn = _norm_key(k)
        if tpl_norm in kn or kn in tpl_norm:
            return k

    # 3) token overlap
    for k in data_dict.keys():
        k_tokens = set([t for t in _norm_key(k).split() if t])
        t_tokens = set([t for t in tpl_norm.split() if t])
        if k_tokens & t_tokens:
            return k

    return None


def infer_fields_schema_mapping(fs, sample_data, resolve_path_fn):
    """Given a fields_schema (possibly with simple types), return a new
    dict where each entry is a mapping dict containing at least 'map_to'
    when a sensible match was found against sample_data.
    """
    if not fs or not isinstance(fs, dict):
        return fs
    out = {}
    for tpl_key, meta in fs.items():
        # existing dict with explicit map_to -> keep
        if isinstance(meta, dict) and meta.get('map_to'):
            out[tpl_key] = meta
            continue

        map_to_candidate = None
        if isinstance(meta, dict):
            map_to_candidate = meta.get('source') or tpl_key
        else:
            map_to_candidate = meta or tpl_key

        matched_key = _find_matching_key(map_to_candidate, tpl_key, sample_data or {}, resolve_path_fn)
        if matched_key:
            out[tpl_key] = {'map_to': matched_key}
        else:
            # keep original type if no mapping found
            out[tpl_key] = meta if isinstance(meta, dict) else meta
    return out


def resolve_path(obj, path):
    """Top-level resolver for dotted/indexed paths to support inference usage."""
    try:
        cur = obj
        if path is None or path == '':
            return None
        import re
        tokens = []
        parts = str(path).split('.')
        for part in parts:
            bracket_parts = re.split(r'\[|\]', part)
            for bp in bracket_parts:
                if bp == '':
                    continue
                tokens.append(bp)
        for tok in tokens:
            if isinstance(cur, list):
                try:
                    idx = int(tok)
                    cur = cur[idx]
                    continue
                except Exception:
                    return None
            if isinstance(cur, dict):
                if tok in cur:
                    cur = cur[tok]
                    continue
                try:
                    cur = cur[int(tok)]
                    continue
                except Exception:
                    return None
            return None
        return cur
    except Exception:
        return None


@csrf_exempt
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
@permission_classes([IsAuthenticated])
def upload_template(request):
    try:
        # Diagnostic info kept via exceptions and return messages; avoid printing
        try:
            auth_header = request.META.get('HTTP_AUTHORIZATION') or (request.headers.get('Authorization') if hasattr(request, 'headers') else None)
        except Exception:
            auth_header = None

        name = request.data.get('name')
        template_type = request.data.get('type')
        html_content = request.data.get('html_content', '')
        html_content = request.data.get('html_content', '')
        # Import HTMLPDFGenerator lazily because WeasyPrint (used by it)
        # may not be installed in the environment (common on Windows).
        try:
            from .services.html_pdf_generator import HTMLPDFGenerator
            default_css = HTMLPDFGenerator.get_default_css()
        except Exception:
            default_css = ''

        css_content = request.data.get('css_content', default_css)
        fields_schema = json.loads(request.data.get('fields_schema', '{}')) if request.data.get('fields_schema') else {}
        template_file = request.FILES.get('template_file')

        # fields_schema is accepted as provided by the client. Do not auto-infer
        # or persist `map_to` entries here; mapping may be applied at render time.

        template = Template.objects.create(
            name=name,
            type=template_type,
            html_content=html_content,
            css_content=css_content,
            fields_schema=fields_schema,
            template_file=template_file
        )

        return Response({
            'id': template.id,
            'name': template.name,
            'message': 'Plantilla creada exitosamente'
        })
    except Exception as e:
        return Response({'error': str(e)}, status=400)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_reports(request):
    reports = Report.objects.all().select_related('maintenance', 'generated_by').values(
        'id', 'title', 'pdf_file', 'generated_at',
        'maintenance__id', 'maintenance__maintenance_type',
        'generated_by__username'
    )
    return Response(list(reports))


@csrf_exempt
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_report(request):
    """Generate a report using the active template and save it to the database."""
    maintenance_id = request.data.get('maintenance_id')
    if not maintenance_id:
        return Response({'error': 'maintenance_id is required'}, status=400)

    try:
        from .models import Maintenance
        maintenance = Maintenance.objects.get(id=int(maintenance_id))
    except Maintenance.DoesNotExist:
        return Response({'error': 'Maintenance not found'}, status=404)

    # Determine which template to use: prefer an explicit `template_id` or
    # `template_name` provided by the frontend. Fall back to the active
    # `ReportTemplate` (legacy) if no Template selected.
    template_obj = None
    active = None
    tpl_id = request.data.get('template_id') or request.data.get('template') or request.data.get('template_name')
    try:
        if tpl_id:
            # allow numeric id or name
            if str(tpl_id).isdigit():
                template_obj = Template.objects.get(id=int(tpl_id))
            else:
                template_obj = Template.objects.filter(name=tpl_id).first()
    except Exception:
        template_obj = None

    if not template_obj:
        active = ReportTemplate.objects.filter(is_active=True).order_by('-updated_at').first()
        if not active:
            return Response({'error': 'No active template found and no template selected'}, status=404)

    # Get data for the maintenance
    try:
        from .services.maintenance_serializer import serialize_maintenance
        data = serialize_maintenance(maintenance.id)
    except Exception:
        return Response({'error': 'Error serializing maintenance data'}, status=500)

    # If the chosen template defines a fields_schema mapping, apply it so
    # template variable names (keys in fields_schema) are resolved from the
    # serialized maintenance data. This allows templates to use friendly
    # names while mapping them to internal maintenance keys (e.g. map_to).
    def _resolve_path(obj, path):
        """Resolve dotted or indexed path in a nested dict/list structure.
        Examples: 'equipment.serial_number', 'activities[0].task', 'nested.arr[1].x'
        Returns None if path cannot be resolved.
        """
        try:
            cur = obj
            if path is None or path == '':
                return None
            # support bracket indexes by converting a[0] to a.0 tokens
            # but keep string keys intact
            # replace [num] with .num
            import re
            tokens = []
            # split by dot first
            parts = path.split('.')
            for part in parts:
                # find bracket parts
                bracket_parts = re.split(r'\[|\]', part)
                for bp in bracket_parts:
                    if bp == '':
                        continue
                    tokens.append(bp)

            for tok in tokens:
                # if current is list-like and tok is integer index
                if isinstance(cur, list):
                    try:
                        idx = int(tok)
                        cur = cur[idx]
                        continue
                    except Exception:
                        return None
                # otherwise treat as dict key
                if isinstance(cur, dict):
                    if tok in cur:
                        cur = cur[tok]
                        continue
                    # try numeric key as fallback
                    try:
                        cur = cur[int(tok)]
                        continue
                    except Exception:
                        return None
                # not indexable
                return None
            return cur
        except Exception:
            return None

    mapped_context = {}
    fs = None
    # Prefer fields_schema on a selected Template instance
    if template_obj is not None:
        fs = getattr(template_obj, 'fields_schema', None)
        html_content = getattr(template_obj, 'html_content', '')
        css_content = getattr(template_obj, 'css_content', '')
    else:
        # Fallback to active ReportTemplate which may not have html content or fields_schema
        fs = getattr(active, 'fields_schema', None)
        html_content = getattr(active, 'html_content', '') if hasattr(active, 'html_content') else ''
        css_content = getattr(active, 'css_content', '') if hasattr(active, 'css_content') else ''

    if fs and isinstance(fs, dict):
        try:
            def _auto_map_value(map_to, tpl_key, data_dict):
                # Try explicit path first
                if map_to:
                    v = _resolve_path(data_dict, map_to) if isinstance(data_dict, (dict, list)) else None
                    if v is not None:
                        return v
                    if isinstance(data_dict, dict) and map_to in data_dict:
                        return data_dict.get(map_to)

                # Heuristic fallback: try to match tpl_key against available keys
                if not isinstance(data_dict, dict):
                    return None

                def _norm(s):
                    try:
                        return ''.join([c for c in str(s).lower() if c.isalnum()])
                    except Exception:
                        return str(s).lower()

                tpl_norm = _norm(tpl_key)
                # 1) exact match ignoring case/characters
                for k in data_dict.keys():
                    if _norm(k) == tpl_norm:
                        return data_dict.get(k)

                # 2) substring match (tpl in key or key in tpl)
                for k in data_dict.keys():
                    kn = _norm(k)
                    if tpl_norm in kn or kn in tpl_norm:
                        return data_dict.get(k)

                # 3) token overlap (split on non-alnum)
                for k in data_dict.keys():
                    k_tokens = set([t for t in _norm(k).split() if t])
                    t_tokens = set([t for t in tpl_norm.split() if t])
                    if k_tokens & t_tokens:
                        return data_dict.get(k)

                return None

            for tpl_key, meta in fs.items():
                map_to = None
                if isinstance(meta, dict):
                    map_to = meta.get('map_to') or meta.get('source') or tpl_key
                else:
                    map_to = meta or tpl_key

                val = _auto_map_value(map_to, tpl_key, data)
                mapped_context[tpl_key] = val
        except Exception:
            mapped_context = {}

    render_context = {**(data or {}), **(mapped_context or {})}

    # Determine template type: prefer template_obj if available, else active
    template_type = 'pdf'
    if template_obj is not None:
        template_type = getattr(template_obj, 'type', 'pdf') or 'pdf'
    elif active is not None:
        template_type = getattr(active, 'type', 'pdf') or 'pdf'

    # Allow clients to request Excel generation using the original .xlsx template
    format_type = request.data.get('format') or request.query_params.get('format') or 'pdf'
    if format_type == 'excel' or template_type == 'excel':
        # Determine which Excel generator to use based on template name or equipment type
        template_name = ''
        if template_obj:
            template_name = (template_obj.name or '').lower()
        
        equipment_type = ''
        if maintenance.equipment:
            equipment_type = (maintenance.equipment.equipment_type or '').lower()
        
        # Use printer/scanner generator if template or equipment indicates it
        is_printer_scanner = (
            'impresora' in template_name or 
            'escaner' in template_name or 
            'scanner' in template_name or
            'printer' in template_name or
            'impresora' in equipment_type or
            'escaner' in equipment_type or
            'scanner' in equipment_type or
            'printer' in equipment_type
        )
        
        try:
            if is_printer_scanner:
                from .services.printer_scanner_excel_generator import PrinterScannerExcelGenerator
                generator = PrinterScannerExcelGenerator()
                filename_prefix = 'rutina_impresora_escaner'
            else:
                from .services.excel_report_generator import ExcelReportGenerator
                generator = ExcelReportGenerator()
                filename_prefix = 'rutina_mantenimiento'
        except Exception as e:
            return Response({'error': f'Generador Excel no disponible: {str(e)}'}, status=501)

        try:
            excel_bytes = generator.generate_report(maintenance)
            # Save to default storage so it appears in reports list
            from django.core.files.base import ContentFile
            from django.core.files.storage import default_storage
            eq_code = maintenance.equipment.code if maintenance.equipment else maintenance.id
            filename = f"reports/{filename_prefix}_{eq_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = default_storage.save(filename, ContentFile(excel_bytes))

            # Create Report record
            try:
                report = Report.objects.create(
                    maintenance=maintenance,
                    title=f"Rutina de Mantenimiento - {maintenance.equipment.name if maintenance.equipment else ''}",
                    content=maintenance.description or '',
                    pdf_file=file_path,
                    generated_by=request.user
                )
            except Exception:
                report = None

            # Return Excel file as blob for frontend download
            response = HttpResponse(excel_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{eq_code}.xlsx"'
            return response
        except Exception as e:
            return Response({'error': str(e)}, status=500)

    # Generate PDF
    if template_type == 'pdf':
        overlays = None
        try:
            overlays_field = request.data.get('_overlays') or request.data.get('overlays')
            if overlays_field:
                if isinstance(overlays_field, str):
                    import json as _json
                    overlays = _json.loads(overlays_field)
                else:
                    overlays = overlays_field
        except Exception:
            overlays = None

        background_bytes = None
        # Try template_obj first, then active for background file
        tpl_with_file = template_obj if template_obj and getattr(template_obj, 'template_file', None) else active
        try:
            if tpl_with_file and tpl_with_file.template_file:
                tpl_with_file.template_file.open('rb')
                background_bytes = tpl_with_file.template_file.read()
                tpl_with_file.template_file.close()
        except Exception:
            background_bytes = None

        # decide which html/css to render (template_obj preferred)
        html_to_use = html_content
        css_to_use = css_content

        if HTMLPDFGenerator:
            pdf_file = HTMLPDFGenerator.render_template(html_to_use, css_to_use, render_context)
        elif ReportLabPDFGenerator:
            pdf_file = ReportLabPDFGenerator.render_template(
                html_to_use,
                css_to_use,
                render_context,
                background_bytes=background_bytes,
                overlays=overlays
            )
        else:
            return Response({'error': 'PDF generation not available'}, status=501)

        # Save the PDF to storage
        from django.core.files.base import ContentFile
        from django.core.files.storage import default_storage
        name_for_file = (template_obj.name if template_obj else (active.name if active else 'report'))
        filename = f"reports/{name_for_file}_{maintenance.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        file_path = default_storage.save(filename, ContentFile(pdf_file.read()))

        # Do not emit debug prints in production; preserve optional saved_size for future logging hooks
        try:
            saved_size = default_storage.size(file_path)
        except Exception:
            saved_size = None

        # Create Report instance
        report = Report.objects.create(
            maintenance=maintenance,
            title=f"Reporte de {maintenance.maintenance_type} - {maintenance.equipment.name if maintenance.equipment else 'Sin equipo'}",
            content='',  # Could add summary
            pdf_file=file_path,
            generated_by=request.user
        )

        # Return PDF as blob for frontend download
        pdf_file.seek(0)
        pdf_content = pdf_file.read()
        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{name_for_file}_{maintenance.id}.pdf"'
        return response
    else:
        return Response({'error': 'Active template is not PDF type'}, status=400)


class ListTemplatesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        templates = Template.objects.all().values('id', 'name', 'type', 'description')
        return Response(list(templates))


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_template(request, template_key):
    try:
        # Allow numeric id or textual key (name)
        if isinstance(template_key, str) and template_key.isdigit():
            template = Template.objects.get(id=int(template_key))
        else:
            template = Template.objects.filter(name=template_key).first()
            if not template:
                # fallback: try exact match on id as int conversion
                try:
                    template = Template.objects.get(id=int(template_key))
                except Exception:
                    template = None

        if not template:
            return Response({'error': 'Plantilla no encontrada'}, status=404)
        template_file_url = None
        try:
            if template.template_file:
                template_file_url = request.build_absolute_uri(template.template_file.url)
        except Exception:
            template_file_url = None

        return Response({
            'id': template.id,
            'name': template.name,
            'type': template.type,
            'description': template.description,
            'html_content': template.html_content,
            'css_content': template.css_content,
            'fields_schema': template.fields_schema,
            'template_file': template_file_url,
        })
    except Template.DoesNotExist:
        return Response({'error': 'Plantilla no encontrada'}, status=404)


@csrf_exempt
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_from_template(request, template_key):
    import logging
    logger = logging.getLogger(__name__)
    try:
        # Diagnostic: log auth header and user info to help debug 403 responses
        auth_header = request.META.get('HTTP_AUTHORIZATION') or request.headers.get('Authorization') if hasattr(request, 'headers') else None
        logger.debug('generate_from_template called: method=%s, auth_header=%s, user=%s, is_authenticated=%s',
                     request.method, auth_header, getattr(request, 'user', None), getattr(request.user, 'is_authenticated', False))
        # Resolve template by id or name
        if isinstance(template_key, str) and template_key.isdigit():
            template = Template.objects.get(id=int(template_key))
        else:
            template = Template.objects.filter(name=template_key).first()
            if not template:
                try:
                    template = Template.objects.get(id=int(template_key))
                except Exception:
                    template = None

        if not template:
            return Response({'error': 'Plantilla no encontrada'}, status=404)
        
        # Get data from request or load from maintenance if maintenance_id is provided
        data = request.data.get('data', {})
        maintenance_id = request.data.get('maintenance_id')
        
        # If maintenance_id is provided, load serialized maintenance data
        if maintenance_id and not data:
            try:
                from .services.maintenance_serializer import serialize_maintenance
                data = serialize_maintenance(int(maintenance_id))
            except Exception as e:
                return Response({'error': f'Error cargando datos del mantenimiento: {str(e)}'}, status=400)

        # Build mapped context from template.fields_schema like in generate_report
        def _resolve_path(obj, path):
            try:
                cur = obj
                if path is None or path == '':
                    return None
                import re
                tokens = []
                parts = str(path).split('.')
                for part in parts:
                    bracket_parts = re.split(r'\[|\]', part)
                    for bp in bracket_parts:
                        if bp == '':
                            continue
                        tokens.append(bp)
                for tok in tokens:
                    if isinstance(cur, list):
                        try:
                            idx = int(tok)
                            cur = cur[idx]
                            continue
                        except Exception:
                            return None
                    if isinstance(cur, dict):
                        if tok in cur:
                            cur = cur[tok]
                            continue
                        try:
                            cur = cur[int(tok)]
                            continue
                        except Exception:
                            return None
                    return None
                return cur
            except Exception:
                return None

        mapped_context = {}
        fs = getattr(template, 'fields_schema', None)
        html_content = getattr(template, 'html_content', '')
        css_content = getattr(template, 'css_content', '')
        if fs and isinstance(fs, dict):
            try:
                def _auto_map_value(map_to, tpl_key, data_dict):
                    # Try explicit path first
                    if map_to:
                        v = _resolve_path(data_dict, map_to) if isinstance(data_dict, (dict, list)) else None
                        if v is not None:
                            return v
                        if isinstance(data_dict, dict) and map_to in data_dict:
                            return data_dict.get(map_to)

                    # Heuristic fallback: try to match tpl_key against available keys
                    if not isinstance(data_dict, dict):
                        return None

                    def _norm(s):
                        try:
                            return ''.join([c for c in str(s).lower() if c.isalnum()])
                        except Exception:
                            return str(s).lower()

                    tpl_norm = _norm(tpl_key)
                    # 1) exact match ignoring case/characters
                    for k in data_dict.keys():
                        if _norm(k) == tpl_norm:
                            return data_dict.get(k)

                    # 2) substring match (tpl in key or key in tpl)
                    for k in data_dict.keys():
                        kn = _norm(k)
                        if tpl_norm in kn or kn in tpl_norm:
                            return data_dict.get(k)

                    # 3) token overlap (split on non-alnum)
                    for k in data_dict.keys():
                        k_tokens = set([t for t in _norm(k).split() if t])
                        t_tokens = set([t for t in tpl_norm.split() if t])
                        if k_tokens & t_tokens:
                            return data_dict.get(k)

                    return None

                for tpl_key, meta in fs.items():
                    map_to = meta.get('map_to') if isinstance(meta, dict) else (meta or tpl_key)
                    val = _auto_map_value(map_to, tpl_key, data)
                    mapped_context[tpl_key] = val
            except Exception:
                mapped_context = {}

        render_context = {**(data or {}), **(mapped_context or {})}

        if template.type == 'pdf':
            # Prefer WeasyPrint-based generator (HTMLPDFGenerator). If no
            # WeasyPrint available, try ReportLab fallback which produces a
            # basic PDF rendering of the template content. If the template
            # has an uploaded `template_file` (image) and the client sends
            # overlay specs, the ReportLab fallback will compose text over
            # the image using percentage coordinates.
            # Extract overlays from request data if present. Expect a list of
            # objects: { text, x_pct, y_pct, font_size }
            overlays = None
            try:
                # overlays may be sent as JSON string or as a JSON body field
                overlays_field = request.data.get('_overlays') or request.data.get('overlays')
                if overlays_field:
                    if isinstance(overlays_field, str):
                        import json as _json
                        overlays = _json.loads(overlays_field)
                    else:
                        overlays = overlays_field
            except Exception:
                overlays = None

            background_bytes = None
            try:
                if template.template_file:
                    # Read the file content (works for local storage or FileStorage)
                    template.template_file.open('rb')
                    background_bytes = template.template_file.read()
                    template.template_file.close()
            except Exception:
                background_bytes = None

            if HTMLPDFGenerator:
                # HTML generator does not support image-overlay in this simple path
                pdf_file = HTMLPDFGenerator.render_template(html_content, css_content, render_context)
            elif ReportLabPDFGenerator:
                pdf_file = ReportLabPDFGenerator.render_template(
                    html_content,
                    css_content,
                    render_context,
                    background_bytes=background_bytes,
                    overlays=overlays
                )
            else:
                return Response({'error': 'Generación de PDF no disponible en este entorno.'}, status=501)
            # Read bytes and validate
            try:
                content = pdf_file.read()
            except Exception:
                return Response({'error': 'Error leyendo el PDF generado'}, status=500)

            if not content or len(content) == 0:
                return Response({'error': 'PDF generado vacío o corrupto'}, status=500)

            # Save the PDF to storage and create a Report record so it appears in "Generados"
            try:
                from django.core.files.base import ContentFile
                from django.core.files.storage import default_storage
                name_for_file = template.name or 'report'
                filename = f"reports/{name_for_file}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                file_path = default_storage.save(filename, ContentFile(content))
            except Exception:
                file_path = None

            # Optionally link to a maintenance if provided in request
            maintenance_obj = None
            try:
                maint_id = request.data.get('maintenance_id')
                if maint_id:
                    from .models import Maintenance
                    maintenance_obj = Maintenance.objects.filter(id=int(maint_id)).first()
            except Exception:
                maintenance_obj = None

            try:
                report = Report.objects.create(
                    maintenance=maintenance_obj,
                    title=f"Reporte {template.name}",
                    content='',
                    pdf_file=file_path or '',
                    generated_by=request.user
                )
            except Exception:
                # If creation fails, continue but don't block download
                report = None

            response = HttpResponse(content, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{template.name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
            # Also include header with saved path/URL when available
            try:
                if file_path:
                    pdf_url = request.build_absolute_uri(default_storage.url(file_path))
                    response['X-Saved-PDF-Url'] = pdf_url
            except Exception:
                pass

            return response
        elif template.type == 'excel':
            # Generate Excel file from template
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import Font, Alignment
                from io import BytesIO
                
                # Load the template file
                if not template.template_file:
                    return Response({'error': 'No hay archivo de plantilla Excel configurado'}, status=400)
                
                # Open the template
                template.template_file.open('rb')
                wb = load_workbook(template.template_file)
                template.template_file.close()
                
                # Get the active sheet
                ws = wb.active
                
                # Fill data based on fields_schema mapping
                # The fields_schema should have cell references like {'A1': {'map_to': 'field_name'}}
                if fs and isinstance(fs, dict):
                    for cell_ref, meta in fs.items():
                        map_to = meta.get('map_to') if isinstance(meta, dict) else (meta or cell_ref)
                        val = mapped_context.get(cell_ref) or render_context.get(map_to)
                        if val is not None:
                            try:
                                ws[cell_ref] = val
                            except Exception:
                                pass
                
                # Save to BytesIO
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                content = output.read()
                
                # Save to storage
                try:
                    from django.core.files.base import ContentFile
                    from django.core.files.storage import default_storage
                    name_for_file = template.name or 'report'
                    filename = f"reports/{name_for_file}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    file_path = default_storage.save(filename, ContentFile(content))
                except Exception:
                    file_path = None
                
                # Create Report record
                maintenance_obj = None
                try:
                    maint_id = request.data.get('maintenance_id')
                    if maint_id:
                        from .models import Maintenance
                        maintenance_obj = Maintenance.objects.filter(id=int(maint_id)).first()
                except Exception:
                    maintenance_obj = None
                
                try:
                    report = Report.objects.create(
                        maintenance=maintenance_obj,
                        title=f"Reporte {template.name}",
                        content='',
                        pdf_file=file_path or '',
                        generated_by=request.user
                    )
                except Exception:
                    report = None
                
                response = HttpResponse(content, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{template.name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
                
                try:
                    if file_path:
                        file_url = request.build_absolute_uri(default_storage.url(file_path))
                        response['X-Saved-File-Url'] = file_url
                except Exception:
                    pass
                
                return response
            except Exception as e:
                return Response({'error': f'Error generando Excel: {str(e)}'}, status=500)
        else:
            return Response({'error': 'Tipo de plantilla no soportado'}, status=400)
    except Template.DoesNotExist:
        return Response({'error': 'Plantilla no encontrada'}, status=404)
    except Exception as e:
        return Response({'error': str(e)}, status=400)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def sample_template_data(request):
    """Devuelve datos de ejemplo serializados para un mantenimiento dado.

    Query params:
      - maintenance_id: int (opcional). Si no se provee, se devuelve el último mantenimiento del usuario.
    """
    maintenance_id = request.query_params.get('maintenance_id')
    try:
        from .services.maintenance_serializer import serialize_maintenance
    except Exception:
        return Response({'error': 'Serializer de mantenimiento no disponible'}, status=500)

    if maintenance_id:
        try:
            data = serialize_maintenance(int(maintenance_id))
            return Response(data)
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    # If no maintenance_id provided, attempt to return most recent maintenance
    try:
        from .models import Maintenance
        m = Maintenance.objects.order_by('-created_at').first()
        if not m:
            return Response({'error': 'No hay mantenimientos disponibles para generar datos de ejemplo'}, status=404)
        data = serialize_maintenance(m.id)
        return Response(data)
    except Exception as e:
        return Response({'error': str(e)}, status=500)


@csrf_exempt
@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_template(request, template_key):
    try:
        if isinstance(template_key, str) and template_key.isdigit():
            template = Template.objects.get(id=int(template_key))
        else:
            template = Template.objects.filter(name=template_key).first()
            if not template:
                try:
                    template = Template.objects.get(id=int(template_key))
                except Exception:
                    template = None

        if not template:
            return Response({'error': 'Plantilla no encontrada'}, status=404)

        template.name = request.data.get('name', template.name)
        template.html_content = request.data.get('html_content', template.html_content)
        template.css_content = request.data.get('css_content', template.css_content)
        fs = request.data.get('fields_schema')
        new_fs = json.loads(fs) if isinstance(fs, str) and fs else (fs or template.fields_schema)
        # Do not infer or persist map_to during update; accept the provided schema as-is.
        template.fields_schema = new_fs
        template.save()
        return Response({'message': 'Plantilla actualizada exitosamente'})
    except Template.DoesNotExist:
        return Response({'error': 'Plantilla no encontrada'}, status=404)
    except Exception as e:
        return Response({'error': str(e)}, status=400)


@csrf_exempt
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_template(request, template_key):
    try:
        if isinstance(template_key, str) and template_key.isdigit():
            template = Template.objects.get(id=int(template_key))
        else:
            template = Template.objects.filter(name=template_key).first()
            if not template:
                try:
                    template = Template.objects.get(id=int(template_key))
                except Exception:
                    template = None

        if not template:
            return Response({'error': 'Plantilla no encontrada'}, status=404)
        template.delete()
        return Response({'message': 'Plantilla eliminada exitosamente'})
    except Template.DoesNotExist:
        return Response({'error': 'Plantilla no encontrada'}, status=404)
    except Exception as e:
        return Response({'error': str(e)}, status=400)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def active_template(request):
    """Return the active report template and the available PDF generator info."""
    active = ReportTemplate.objects.filter(is_active=True).order_by('-updated_at').first()
    generator = 'reportlab'
    try:
        from .services.html_pdf_generator import HTMLPDFGenerator
        if HTMLPDFGenerator:
            generator = 'weasyprint'
    except Exception:
        generator = 'reportlab'

    if not active:
        return Response({'template': None, 'generator': generator})

    return Response({
        'template': {
            'id': active.id,
            'name': active.name,
            'description': active.description,
            'is_active': active.is_active,
        },
        'generator': generator
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def suggest_mappings(request, template_key):
    """Return suggested fields_schema mappings for a template without persisting.

    Request body (JSON):
      - maintenance_id: optional int to use a specific maintenance sample
      - data: optional dict to use as sample data directly

    Response: {
      'suggested_fields_schema': { tpl_key: { 'map_to': suggested_key } | original_meta },
      'sample_data_keys': [ ... ]
    }
    """
    try:
        # Resolve template by id or name
        if isinstance(template_key, str) and template_key.isdigit():
            template = Template.objects.get(id=int(template_key))
        else:
            template = Template.objects.filter(name=template_key).first()
            if not template:
                try:
                    template = Template.objects.get(id=int(template_key))
                except Exception:
                    template = None

        if not template:
            return Response({'error': 'Plantilla no encontrada'}, status=404)

        # Obtain sample data
        sample = None
        if request.data.get('data'):
            sample = request.data.get('data')
        else:
            maintenance_id = request.data.get('maintenance_id')
            try:
                from .services.maintenance_serializer import serialize_maintenance
            except Exception:
                return Response({'error': 'Serializer de mantenimiento no disponible'}, status=500)

            if maintenance_id:
                try:
                    sample = serialize_maintenance(int(maintenance_id))
                except Exception as e:
                    return Response({'error': str(e)}, status=400)
            else:
                try:
                    from .models import Maintenance
                    m = Maintenance.objects.order_by('-created_at').first()
                    if not m:
                        return Response({'error': 'No hay mantenimientos disponibles para generar datos de ejemplo'}, status=404)
                    sample = serialize_maintenance(m.id)
                except Exception as e:
                    return Response({'error': str(e)}, status=500)

        # Use helper to infer mappings (non-persistent)
        fs = getattr(template, 'fields_schema', None) or {}
        suggestions = infer_fields_schema_mapping(fs, sample or {}, resolve_path)

        sample_keys = list(sample.keys()) if isinstance(sample, dict) else []
        return Response({'suggested_fields_schema': suggestions, 'sample_data_keys': sample_keys})
    except Template.DoesNotExist:
        return Response({'error': 'Plantilla no encontrada'}, status=404)
    except Exception as e:
        return Response({'error': str(e)}, status=400)
