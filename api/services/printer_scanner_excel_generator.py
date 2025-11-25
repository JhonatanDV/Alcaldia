# -*- coding: utf-8 -*-
"""
Generador de reportes Excel para mantenimientos de impresoras y escáneres.
Usa la plantilla original y rellena las zonas marcadas con datos del mantenimiento.
"""
import os
from datetime import datetime
from io import BytesIO
from PIL import Image

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from django.conf import settings


class PrinterScannerExcelGenerator:
    """
    Genera reportes Excel para mantenimientos de impresoras y escáneres.
    """

    # Ruta a la plantilla original
    TEMPLATE_PATH = os.path.join(
        settings.BASE_DIR, 
        'plantillas',
        'Mantenimiento Escaneres e impresoras',
        'rutina_mantenimiento_preventivo_impresoras_y_escaner_v2.xlsx'
    )

    # Actividades de ESCÁNER (filas 15-16)
    SCANNER_ACTIVITIES = {
        15: {
            'left': 'Limpieza general',
            'right': 'Configuracion del Equipo'
        },
        16: {
            'left': 'Alineacion de Papel',
            'right': 'Pruebas de funcionamiento'
        },
    }

    # Actividades de IMPRESORAS (filas 21-28)
    PRINTER_ACTIVITIES = {
        21: {'left': 'Limpiéza de carcaza', 'right': 'Limpieza de engranaje'},
        22: {'left': 'Limpieza toner', 'right': 'Limpieza de fusor o rodillo'},
        23: {'left': 'Limpiéza tarjeta logica', 'right': 'Limpieza tarjeta de poder'},
        24: {'left': 'Limpiéza de sensores', 'right': 'Alineacion de  rodillos alimentación de'},
        25: {'left': 'Limpiéza de carcaza', 'right': 'Configuracion del Equipo'},
        26: {'left': 'Limpieza de correas dentadas o guias', 'right': 'Pruebas de funcionamiento'},
        27: {'left': 'Limpieza de Ventiladores', 'right': None},
        28: {'left': 'Limpieza de cabezal impresora matriz de', 'right': None},
    }

    # Alias para actividades
    ACTIVITY_ALIASES = {
        'limpieza general': ['limpieza_general', 'limpieza', 'Limpieza general'],
        'configuracion del equipo': ['configuracion', 'config', 'Configuracion del Equipo'],
        'alineacion de papel': ['alineacion', 'papel', 'Alineacion de Papel'],
        'pruebas de funcionamiento': ['pruebas', 'funcionamiento', 'Pruebas de funcionamiento'],
        'limpieza de carcaza': ['carcaza', 'Limpiéza de carcaza', 'limpieza_carcaza'],
        'limpieza de engranaje': ['engranaje', 'Limpieza de engranaje', 'limpieza_engranaje'],
        'limpieza toner': ['toner', 'Limpieza toner', 'limpieza_toner'],
        'limpieza de fusor o rodillo': ['fusor', 'rodillo', 'Limpieza de fusor o rodillo'],
        'limpieza tarjeta logica': ['tarjeta_logica', 'Limpiéza tarjeta logica'],
        'limpieza tarjeta de poder': ['tarjeta_poder', 'Limpieza tarjeta de poder'],
        'limpieza de sensores': ['sensores', 'Limpiéza de sensores'],
        'alineacion de  rodillos alimentación de': ['rodillos', 'alimentacion', 'Alineacion de  rodillos alimentación de'],
        'limpieza de correas dentadas o guias': ['correas', 'guias', 'Limpieza de correas dentadas o guias'],
        'limpieza de ventiladores': ['ventiladores', 'Limpieza de Ventiladores'],
        'limpieza de cabezal impresora matriz de': ['cabezal', 'matriz', 'Limpieza de cabezal impresora matriz de'],
    }

    def __init__(self):
        if not os.path.exists(self.TEMPLATE_PATH):
            raise FileNotFoundError(f"Plantilla no encontrada: {self.TEMPLATE_PATH}")

    def generate_report(self, maintenance) -> bytes:
        """
        Genera el reporte Excel para un mantenimiento de impresora/escáner.
        
        Args:
            maintenance: Objeto Maintenance
            
        Returns:
            bytes: Contenido del archivo Excel generado
        """
        wb = openpyxl.load_workbook(self.TEMPLATE_PATH)
        ws = wb.active

        # Rellenar encabezado
        self._fill_header(ws, maintenance)

        # Rellenar marca y modelo
        self._fill_equipment_info(ws, maintenance)

        # Rellenar actividades según tipo de equipo
        equipment_type = maintenance.equipment.equipment_type if hasattr(maintenance.equipment, 'equipment_type') else ''
        
        if 'scanner' in equipment_type.lower() or 'escaner' in equipment_type.lower():
            self._fill_scanner_activities(ws, maintenance)
        elif 'printer' in equipment_type.lower() or 'impresora' in equipment_type.lower():
            self._fill_printer_activities(ws, maintenance)
        else:
            # Rellenar ambas si no está claro
            self._fill_scanner_activities(ws, maintenance)
            self._fill_printer_activities(ws, maintenance)

        # Rellenar observaciones y firmas
        self._fill_observations(ws, maintenance)
        self._fill_signatures(ws, maintenance)

        # Embeber fotos al final
        self._embed_maintenance_photos(ws, maintenance)

        # Calificación del servicio
        self._fill_service_rating(ws, maintenance)

        # Guardar a bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def _fill_header(self, ws, maintenance):
        """Rellena encabezado (filas 9-10)."""
        # Sede (A9)
        sede = maintenance.sede or (maintenance.sede_rel.nombre if maintenance.sede_rel else '') or ''
        ws['A9'] = f"Sede: {sede}"

        # Dependencia (C9)
        dependencia = maintenance.dependencia or (maintenance.dependencia_rel.nombre if maintenance.dependencia_rel else '') or ''
        ws['C9'] = f"Dependencia: {dependencia}"

        # Oficina (G9)
        oficina = maintenance.oficina or maintenance.ubicacion or ''
        ws['G9'] = f"    Oficina: {oficina}"

        # Placa (A10)
        placa = maintenance.placa or maintenance.equipment.code or ''
        ws['A10'] = f"Placa: {placa}"

        # Fecha (C10)
        fecha = maintenance.maintenance_date
        if fecha:
            fecha_str = fecha.strftime('%d/%m/%Y')
            ws['C10'] = f"Fecha Mantenimiento: {fecha_str}"

        # Horas (G10)
        hora_inicio = maintenance.hora_inicio.strftime('%H:%M') if maintenance.hora_inicio else '____'
        hora_final = maintenance.hora_final.strftime('%H:%M') if maintenance.hora_final else '____'
        ws['G10'] = f"Hora Inicio {hora_inicio}               Hora Final {hora_final}"

    def _fill_equipment_info(self, ws, maintenance):
        """Rellena marca y modelo (fila 11-12)."""
        equipment = maintenance.equipment
        
        # Marca y modelo impresora (A11)
        marca_modelo = f"{equipment.brand or ''} {equipment.model or ''}".strip()
        ws['A11'] = f"Marca y Modelo Impresora : {marca_modelo}"

        # Marca y modelo escáner (D11) - puede ser el mismo
        ws['D11'] = f"Marca y Modelo Escaner : {marca_modelo}"

    def _fill_scanner_activities(self, ws, maintenance):
        """Rellena actividades de escáner (filas 15-16)."""
        activities = maintenance.activities or {}
        
        for row, act_info in self.SCANNER_ACTIVITIES.items():
            if act_info['left']:
                si_val, na_val = self._get_activity_value(activities, act_info['left'])
                ws.cell(row=row, column=4).value = si_val  # D
                ws.cell(row=row, column=5).value = na_val  # E
            
            if act_info.get('right'):
                si_val, na_val = self._get_activity_value(activities, act_info['right'])
                ws.cell(row=row, column=9).value = si_val   # I
                ws.cell(row=row, column=10).value = na_val  # J

    def _fill_printer_activities(self, ws, maintenance):
        """Rellena actividades de impresoras (filas 21-28)."""
        activities = maintenance.activities or {}
        
        for row, act_info in self.PRINTER_ACTIVITIES.items():
            if act_info['left']:
                si_val, na_val = self._get_activity_value(activities, act_info['left'])
                ws.cell(row=row, column=4).value = si_val  # D
                ws.cell(row=row, column=5).value = na_val  # E
            
            if act_info.get('right'):
                si_val, na_val = self._get_activity_value(activities, act_info['right'])
                ws.cell(row=row, column=9).value = si_val   # I
                ws.cell(row=row, column=10).value = na_val  # J

    def _get_activity_value(self, activities: dict, activity_name: str) -> tuple:
        """Busca el valor de una actividad y retorna (si_value, na_value)."""
        if not activities:
            return ('', '')

        activity_lower = activity_name.lower().strip()
        
        # Buscar por nombre exacto
        for key, value in activities.items():
            if key.lower().strip() == activity_lower:
                return self._parse_activity_value(value)
        
        # Buscar por alias
        aliases = self.ACTIVITY_ALIASES.get(activity_lower, [])
        for alias in aliases:
            for key, value in activities.items():
                if key.lower().strip() == alias.lower():
                    return self._parse_activity_value(value)

        # Buscar coincidencia parcial
        for key, value in activities.items():
            if activity_lower in key.lower() or key.lower() in activity_lower:
                return self._parse_activity_value(value)
        
        return ('', '')

    def _parse_activity_value(self, value) -> tuple:
        """Parsea valor de actividad y retorna (si, na)."""
        if value is None:
            return ('', '')
        
        if isinstance(value, bool):
            return ('■', '') if value else ('', '■')
        
        if isinstance(value, str):
            val_lower = value.lower().strip()
            if val_lower in ('si', 'sí', 'yes', 'true', '1', 'x', '■'):
                return ('■', '')
            elif val_lower in ('no', 'false', '0'):
                return ('', '■')
            elif val_lower in ('na', 'n.a', 'n/a', 'n.a.'):
                return ('', '■')
        
        if isinstance(value, dict):
            si = '■' if value.get('si') or value.get('SI') else ''
            na = '■' if value.get('na') or value.get('NA') or value.get('n.a') else ''
            return (si, na)
        
        return ('', '')

    def _fill_observations(self, ws, maintenance):
        """Rellena observaciones."""
        # Observaciones generales - A35 es la celda top-left del rango merged A35:J37
        obs = maintenance.observaciones_generales or maintenance.observations or ''
        if obs:
            # Escribir en la celda top-left del rango merged
            ws.cell(35, 1).value = obs

    def _fill_signatures(self, ws, maintenance):
        """Rellena firmas (filas 39-42)."""
        # Responsable (A39)
        nombre_tecnico = maintenance.performed_by or maintenance.elaborado_por or ''
        ws['A39'] = f"Responsable Mantenimiento: {nombre_tecnico}"

        # Usuario (F39)
        nombre_usuario = ''
        if maintenance.second_signature:
            nombre_usuario = maintenance.second_signature.signer_name or ''
        elif maintenance.revisado_por:
            nombre_usuario = maintenance.revisado_por
        ws['F39'] = f"Usuario del Equipo:{nombre_usuario}"

        # Firma técnico en A40
        signature = maintenance.signature
        if signature and signature.image:
            try:
                self._embed_image(ws, signature.image, 'A40', max_width=150, max_height=60)
            except Exception as e:
                print(f"No se pudo embeber firma técnico: {e}")

        # Firma usuario en F40
        second_signature = maintenance.second_signature
        if second_signature and second_signature.image:
            try:
                self._embed_image(ws, second_signature.image, 'F40', max_width=150, max_height=60)
            except Exception as e:
                print(f"No se pudo embeber firma usuario: {e}")

        # Nombre técnico (A41)
        ws['A41'] = f"Nombre: {nombre_tecnico}"

        # Cargo técnico (A42)
        cargo = 'Técnico de Mantenimiento'
        if maintenance.technician:
            cargo = getattr(maintenance.technician, 'cargo', cargo) or cargo
        ws['A42'] = f"Cargo: {cargo}"

        # Nombre usuario (F41)
        ws['F41'] = f"Nombre: {nombre_usuario}"

        # Cédula usuario (F42)
        cedula = ''
        ws['F42'] = f"Cedula: {cedula}"

    def _embed_maintenance_photos(self, ws, maintenance):
        """Embebe fotos al final del documento."""
        photos = maintenance.photos.all()
        if not photos:
            return
        
        start_row = 50
        row_spacing = 12
        
        for idx, photo in enumerate(photos):
            if photo.image:
                try:
                    anchor_row = start_row + (idx * row_spacing)
                    anchor_cell = f'A{anchor_row}'
                    self._embed_image(ws, photo.image, anchor_cell, max_width=200, max_height=150)
                    
                    if photo.caption:
                        desc_row = anchor_row + 8
                        ws.cell(desc_row, 1).value = f"Foto {idx+1}: {photo.caption}"
                except Exception as e:
                    print(f"No se pudo embeber foto {idx+1}: {e}")

    def _fill_service_rating(self, ws, maintenance):
        """Rellena calificación del servicio."""
        calificacion = (maintenance.calificacion_servicio or '').lower().strip()
        
        # Las celdas de calificación pueden estar en filas similares a la otra plantilla
        # Ajustar según la plantilla real
        pass

    def _embed_image(self, ws, image_field, anchor_cell: str, max_width: int = 150, max_height: int = 100):
        """Embebe una imagen en la hoja."""
        try:
            if hasattr(image_field, 'path'):
                img_path = image_field.path
                if os.path.exists(img_path):
                    pil_img = Image.open(img_path)
                else:
                    img_bytes = image_field.read()
                    pil_img = Image.open(BytesIO(img_bytes))
            else:
                img_bytes = image_field.read()
                pil_img = Image.open(BytesIO(img_bytes))
            
            pil_img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
            
            img_buffer = BytesIO()
            pil_img.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            xl_img = XLImage(img_buffer)
            xl_img.anchor = anchor_cell
            ws.add_image(xl_img)
            
        except Exception as e:
            print(f"Error embebiendo imagen en {anchor_cell}: {e}")
            raise


def generate_printer_scanner_report(maintenance) -> bytes:
    """
    Función de conveniencia para generar reporte de impresora/escáner.
    """
    generator = PrinterScannerExcelGenerator()
    return generator.generate_report(maintenance)
