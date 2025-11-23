from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO


class ExcelGenerator:
    def generate(self, data: dict) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Reporte Mantenimiento'

        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        ws['A1'] = 'REPORTE DE MANTENIMIENTO'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        row = 3
        fields = [
            ('Código:', 'codigo'),
            ('Equipo:', 'equipment_name'),
            ('Código Equipo:', 'equipment_code'),
            ('Serial:', 'equipment_serial'),
            ('Tipo:', 'maintenance_type'),
            ('Estado:', 'status'),
            ('Fecha Programada:', 'scheduled_date'),
            ('Fecha Completado:', 'completion_date'),
            ('Técnico:', 'technician_name'),
            ('Sede:', 'sede'),
            ('Dependencia:', 'dependencia'),
        ]

        for label, key in fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = data.get(key, '')
            row += 1

        activities = data.get('activities', []) or []
        if activities:
            row += 2
            ws[f'A{row}'] = 'ACTIVIDADES'
            ws[f'A{row}'].font = Font(size=14, bold=True)
            row += 1

            ws[f'A{row}'] = '#'
            ws[f'B{row}'] = 'Descripción'
            ws[f'C{row}'] = 'Estado'
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].fill = header_fill
                ws[f'{col}{row}'].font = header_font
            row += 1

            for idx, act in enumerate(activities, 1):
                ws[f'A{row}'] = idx
                ws[f'B{row}'] = act.get('description', 'N/A') if isinstance(act, dict) else str(act)
                ws[f'C{row}'] = act.get('status', 'N/A') if isinstance(act, dict) else ''
                row += 1

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
