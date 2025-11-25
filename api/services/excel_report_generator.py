# -*- coding: utf-8 -*-
"""
Generador de reportes Excel para mantenimientos de equipos de cómputo.
Usa la plantilla original y rellena las zonas marcadas con ■ con datos del mantenimiento.
"""
import os
import copy
from datetime import datetime
from io import BytesIO
from PIL import Image

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from django.conf import settings
from django.core.files.storage import default_storage


class ExcelReportGenerator:
    """
    Genera reportes Excel a partir de la plantilla original de rutina de mantenimiento.
    Mantiene el formato exacto y rellena solo las zonas marcadas.
    """

    # Ruta a la plantilla original
    TEMPLATE_PATH = os.path.join(
        settings.BASE_DIR, 
        'plantillas', 
        'rutina_mantenimiento_preventivo_equipos_de_computo_.xlsx'
    )

    # Mapeo de actividades de HARDWARE (filas 11-14) con su posición en la plantilla
    # Columna A=actividad izquierda, D=SI, E=N.A
    # Columna F=actividad derecha, I=SI, J=N.A
    HARDWARE_ACTIVITIES = {
        # Fila 11
        11: {
            'left': 'Limpieza interna de la torre',
            'right': 'Ajuste de tarjetas (Memoria - Video - Red)'
        },
        # Fila 12
        12: {
            'left': 'Limpieza del teclado',
            'right': 'Lubricación del ventilador de la torre'
        },
        # Fila 13
        13: {
            'left': 'Limpieza del monitor',
            'right': 'Lubricación del ventilador de la fuente'
        },
        # Fila 14
        14: {
            'left': 'Verificación de cables de poder y de datos',
            'right': 'Lubricación del ventilador del procesador'
        },
    }

    # Mapeo de actividades de SOFTWARE (filas 18-30)
    SOFTWARE_ACTIVITIES = {
        18: {'left': 'Crear partición de datos', 'right': 'Desactivar aplicaciones al inicio de Windows'},
        19: {'left': 'Mover información a partición de datos', 'right': 'Configurar página de inicio navegador'},
        20: {'left': 'Reinstalar sistema operativo', 'right': 'Configurar fondo de pantalla institucional'},
        21: {'left': 'Instalar antivirus', 'right': 'Configurar protector de pantalla institucional'},
        22: {'left': 'Análisis en busca de software malicioso', 'right': 'Verificar funcionamiento general'},
        23: {'left': 'Diagnosticar funcionamiento aplicaciones instaladas', 'right': 'Inventario de equipo'},
        24: {'left': 'Suspender actualizaciones automáticas S.O.', 'right': 'Limpieza de registros y eliminación de archivos temporales'},
        25: {'left': 'Instalar programas esenciales (ofimática, grabador de discos)', 'right': 'Creación Punto de Restauración'},
        26: {'left': 'Configurar usuarios administrador local', 'right': 'Verificar espacio en disco'},
        27: {'left': 'Modificar contraseña de administrador', 'right': 'Desactivar software no autorizado'},
        28: {'left': 'Configurar nombre equipo', 'right': 'Analizar disco duro'},
        29: {'left': 'El equipo tiene estabilizador', 'right': 'El usuario de Windows tiene contraseña'},
        30: {'left': 'El escritorio está limpio', 'right': None},
    }

    # Alias para nombres de actividades (para hacer match con activities JSON)
    ACTIVITY_ALIASES = {
        # Hardware activities
        'limpieza interna de la torre': ['limpieza_interna_torre', 'limpieza interna', 'limpieza', 'Limpieza interna de la torre'],
        'limpieza del teclado': ['limpieza_teclado', 'teclado', 'Limpieza del teclado'],
        'limpieza del monitor': ['limpieza_monitor', 'monitor', 'Limpieza del monitor'],
        'verificación de cables de poder y de datos': ['verificacion_cables', 'cables', 'revision', 'Verificación de cables de poder y de datos'],
        'ajuste de tarjetas (memoria - video - red)': ['ajuste_tarjetas', 'tarjetas', 'Ajuste de tarjetas (Memoria - Video - Red)'],
        'lubricación del ventilador de la torre': ['lubricacion_ventilador_torre', 'ventilador torre', 'Lubricación del ventilador de la torre'],
        'lubricación del ventilador de la fuente': ['lubricacion_ventilador_fuente', 'ventilador fuente', 'Lubricación del ventilador de la fuente'],
        'lubricación del ventilador del procesador': ['lubricacion_ventilador_procesador', 'ventilador procesador', 'Lubricación del ventilador del procesador'],
        # Software activities
        'crear partición de datos': ['crear_particion', 'particion', 'Crear partición de datos'],
        'mover información a partición de datos': ['mover_informacion', 'mover particion', 'Mover información a partición de datos'],
        'reinstalar sistema operativo': ['reinstalar_so', 'reinstalar', 'Reinstalar sistema operativo'],
        'instalar antivirus': ['instalar_antivirus', 'antivirus', 'Instalar antivirus'],
        'análisis en busca de software malicioso': ['analisis_malware', 'malware', 'Análisis en busca de software malicioso'],
        'diagnosticar funcionamiento aplicaciones instaladas': ['diagnosticar_apps', 'diagnosticar', 'Diagnosticar funcionamiento aplicaciones instaladas'],
        'suspender actualizaciones automáticas s.o.': ['suspender_actualizaciones', 'actualizaciones', 'actualizacion', 'Suspender actualizaciones automáticas S.O.'],
        'instalar programas esenciales (ofimática, grabador de discos)': ['instalar_programas', 'programas esenciales', 'Instalar programas esenciales (ofimática, grabador de discos)'],
        'configurar usuarios administrador local': ['configurar_admin', 'admin local', 'Configurar usuarios administrador local'],
        'modificar contraseña de administrador': ['modificar_password', 'password admin', 'Modificar contraseña de administrador'],
        'configurar nombre equipo': ['configurar_nombre', 'nombre equipo', 'Configurar nombre equipo'],
        'el equipo tiene estabilizador': ['tiene_estabilizador', 'estabilizador', 'El equipo tiene estabilizador'],
        'el escritorio esta limpio': ['escritorio_limpio', 'escritorio', 'El escritorio esta limpio'],
        'desactivar aplicaciones al inicio de windows': ['desactivar_inicio', 'inicio windows', 'Desactivar aplicaciones al inicio de Windows'],
        'configurar página de inicio navegador': ['configurar_navegador', 'navegador', 'Configurar página de inicio navegador'],
        'configurar fondo de pantalla institucional': ['fondo_pantalla', 'fondo', 'Configurar fondo de pantalla institucional'],
        'configurar protector de pantalla institucional': ['protector_pantalla', 'protector', 'Configurar protector de pantalla institucional'],
        'verificar funcionamiento general': ['verificar_funcionamiento', 'funcionamiento', 'Verificar funcionamiento general'],
        'inventario de equipo': ['inventario', 'inventario equipo', 'Inventario de equipo'],
        'limpieza de registros y eliminación de archivos temporales': ['limpieza_registros', 'archivos temporales', 'Limpieza de registros y eliminación de archivos temporales'],
        'creación punto de restauración': ['punto_restauracion', 'restauracion', 'Creación Punto de Restauración'],
        'verificar espacio en disco': ['verificar_disco', 'espacio disco', 'Verificar espacio en disco'],
        'desactivar software no autorizado': ['desactivar_software', 'software no autorizado', 'Desactivar software no autorizado'],
        'analizar disco duro': ['analizar_disco', 'disco duro', 'Analizar disco duro'],
        'el usuario de windows tiene contraseña': ['usuario_password', 'password usuario', 'EL Usuario de windows tiene contraseña'],
    }
    
    # Mapeo genérico: si el JSON tiene claves genéricas, aplicar a múltiples actividades
    GENERIC_ACTIVITIES_MAP = {
        'limpieza': [
            'Limpieza interna de la torre',
            'Limpieza del teclado', 
            'Limpieza del monitor',
        ],
        'revision': [
            'Verificación de cables de poder y de datos',
            'Verificar funcionamiento general',
        ],
        'actualizacion': [
            'Suspender actualizaciones automáticas S.O.',
        ],
    }

    def __init__(self):
        if not os.path.exists(self.TEMPLATE_PATH):
            raise FileNotFoundError(f"Plantilla no encontrada: {self.TEMPLATE_PATH}")

    def generate_report(self, maintenance) -> bytes:
        """
        Genera el reporte Excel para un mantenimiento específico.
        
        Args:
            maintenance: Objeto Maintenance con los datos a rellenar
            
        Returns:
            bytes: Contenido del archivo Excel generado
        """
        # Cargar plantilla original (preservar formato)
        wb = openpyxl.load_workbook(self.TEMPLATE_PATH)
        ws = wb.active

        # Rellenar datos del encabezado
        self._fill_header(ws, maintenance)

        # Rellenar actividades de hardware
        self._fill_hardware_activities(ws, maintenance)

        # Rellenar actividades de software
        self._fill_software_activities(ws, maintenance)

        # Rellenar observaciones
        self._fill_observations(ws, maintenance)

        # Rellenar firmas (incluye embeber imágenes)
        self._fill_signatures(ws, maintenance)

        # Embeber fotos de mantenimiento
        self._embed_maintenance_photos(ws, maintenance)

        # Rellenar calificación del servicio
        self._fill_service_rating(ws, maintenance)

        # Rellenar observaciones del usuario
        self._fill_user_observations(ws, maintenance)

        # Guardar a bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def _fill_header(self, ws, maintenance):
        """Rellena los datos del encabezado del formulario."""
        equipment = maintenance.equipment

        # Sede (A7)
        sede = maintenance.sede or (maintenance.sede_rel.nombre if maintenance.sede_rel else '') or ''
        ws['A7'] = f"Sede: {sede}"

        # Dependencia (C7)
        dependencia = maintenance.dependencia or (maintenance.dependencia_rel.nombre if maintenance.dependencia_rel else '') or ''
        ws['C7'] = f"Dependencia: {dependencia}"

        # Oficina (H7)
        oficina = maintenance.oficina or maintenance.ubicacion or ''
        ws['H7'] = f"          Oficina:{oficina}"

        # Placa Torre (A8)
        placa = maintenance.placa or equipment.code or ''
        ws['A8'] = f"Placa Torre: (Si no tiene informar\npara su identificación):  {placa}"

        # Fecha Mantenimiento (C8) - Día, Mes, Año
        fecha = maintenance.maintenance_date
        if fecha:
            dia = fecha.day
            mes = fecha.strftime('%B') if hasattr(fecha, 'strftime') else str(fecha.month)
            anio = fecha.year
            ws['C8'] = f"Fecha \nMantenimiento:  Día  {dia}     Mes      {mes}               Año  {anio}        "
        
        # Hora Inicio y Final (H8)
        hora_inicio = maintenance.hora_inicio.strftime('%H:%M') if maintenance.hora_inicio else '____'
        hora_final = maintenance.hora_final.strftime('%H:%M') if maintenance.hora_final else '____'
        ws['H8'] = f"Hora Inicio       {hora_inicio}       Hora Final        {hora_final}      "

    def _get_activity_value(self, activities: dict, activity_name: str) -> tuple:
        """
        Busca el valor de una actividad en el diccionario de actividades.
        
        Returns:
            tuple: (si_value, na_value) donde cada uno es '■' o ''
        """
        if not activities:
            return ('', '')

        # Normalizar nombre de actividad
        activity_lower = activity_name.lower().strip()
        
        # Buscar por nombre exacto
        for key, value in activities.items():
            key_lower = key.lower().strip()
            if key_lower == activity_lower:
                return self._parse_activity_value(value)
        
        # Buscar por alias
        aliases = self.ACTIVITY_ALIASES.get(activity_lower, [])
        for alias in aliases:
            alias_lower = alias.lower()
            for key, value in activities.items():
                key_lower = key.lower().strip()
                if key_lower == alias_lower or alias_lower in key_lower:
                    return self._parse_activity_value(value)

        # Buscar por coincidencia parcial
        for key, value in activities.items():
            key_lower = key.lower().strip()
            # Si la actividad contiene palabras clave del nombre
            words = activity_lower.split()
            if len(words) >= 2:
                if all(w in key_lower for w in words[:2]):
                    return self._parse_activity_value(value)
        
        # Buscar en mapeo genérico (p.ej. 'limpieza' aplica a varias actividades)
        for generic_key, target_activities in self.GENERIC_ACTIVITIES_MAP.items():
            # Si activity_name está en la lista de actividades afectadas por la clave genérica
            for target in target_activities:
                if target.lower() == activity_lower or activity_lower in target.lower():
                    # Buscar el valor de la clave genérica en activities
                    for key, value in activities.items():
                        if key.lower() == generic_key.lower():
                            return self._parse_activity_value(value)

        return ('', '')

    def _parse_activity_value(self, value) -> tuple:
        """
        Parsea el valor de una actividad y devuelve (si, na).
        
        El valor puede ser:
        - True/False/'si'/'no' -> (■, '') o ('', ■)
        - 'na'/'n.a' -> ('', ■)
        - dict con 'si' y 'na' -> correspondiente
        """
        if value is None:
            return ('', '')
        
        if isinstance(value, bool):
            return ('■', '') if value else ('', '■')
        
        if isinstance(value, str):
            val_lower = value.lower().strip()
            if val_lower in ('si', 'sí', 'yes', 'true', '1', 'x', '■'):
                return ('■', '')
            elif val_lower in ('no', 'false', '0'):
                return ('', '■')
            elif val_lower in ('na', 'n.a', 'n/a', 'n.a.'):
                return ('', '■')
        
        if isinstance(value, dict):
            si = '■' if value.get('si') or value.get('SI') else ''
            na = '■' if value.get('na') or value.get('NA') or value.get('n.a') else ''
            return (si, na)
        
        return ('', '')

    def _fill_hardware_activities(self, ws, maintenance):
        """Rellena las actividades de hardware (filas 11-14)."""
        activities = maintenance.activities or {}
        
        for row, act_info in self.HARDWARE_ACTIVITIES.items():
            # Actividad izquierda (columnas D=SI, E=N.A)
            if act_info['left']:
                si_val, na_val = self._get_activity_value(activities, act_info['left'])
                ws.cell(row=row, column=4).value = si_val  # D
                ws.cell(row=row, column=5).value = na_val  # E
            
            # Actividad derecha (columnas I=SI, J=N.A)
            if act_info['right']:
                si_val, na_val = self._get_activity_value(activities, act_info['right'])
                ws.cell(row=row, column=9).value = si_val   # I
                ws.cell(row=row, column=10).value = na_val  # J

    def _fill_software_activities(self, ws, maintenance):
        """Rellena las actividades de software (filas 18-30)."""
        activities = maintenance.activities or {}
        
        for row, act_info in self.SOFTWARE_ACTIVITIES.items():
            # Actividad izquierda (columnas D=SI, E=N.A/NO)
            if act_info['left']:
                si_val, na_val = self._get_activity_value(activities, act_info['left'])
                ws.cell(row=row, column=4).value = si_val  # D
                ws.cell(row=row, column=5).value = na_val  # E
            
            # Actividad derecha (columnas I=SI, J=N.A/NO)
            if act_info.get('right'):
                si_val, na_val = self._get_activity_value(activities, act_info['right'])
                ws.cell(row=row, column=9).value = si_val   # I
                ws.cell(row=row, column=10).value = na_val  # J

    def _fill_observations(self, ws, maintenance):
        """Rellena las observaciones generales y de seguridad."""
        # Observaciones generales (A32)
        obs_generales = maintenance.observaciones_generales or maintenance.observations or ''
        ws['A32'] = obs_generales

        # Observaciones seguridad de la información (A34)
        obs_seguridad = maintenance.observaciones_seguridad or ''
        ws['A34'] = obs_seguridad

    def _fill_signatures(self, ws, maintenance):
        """Rellena los nombres de las firmas y embebe imágenes de firma."""
        # Responsable mantenimiento - Nombre (A38)
        nombre_tecnico = maintenance.performed_by or maintenance.elaborado_por or ''
        ws['A38'] = f"Nombre: {nombre_tecnico}"

        # Cargo del técnico (A39)
        cargo = 'Técnico de Mantenimiento'
        if maintenance.technician:
            # Intentar obtener cargo del perfil si existe
            cargo = getattr(maintenance.technician, 'cargo', cargo) or cargo
        ws['A39'] = f"Cargo: {cargo}"

        # Embeber firma del técnico en A37
        signature = maintenance.signature
        if signature and signature.image:
            try:
                self._embed_image(ws, signature.image, 'A37', max_width=150, max_height=60)
            except Exception as e:
                print(f"No se pudo embeber firma técnico: {e}")

        # Usuario del equipo - Nombre (F38)
        nombre_usuario = ''
        if maintenance.second_signature:
            nombre_usuario = maintenance.second_signature.signer_name or ''
        elif maintenance.revisado_por:
            nombre_usuario = maintenance.revisado_por
        ws['F38'] = f"Nombre: {nombre_usuario}"

        # Cédula usuario (F39)
        cedula = ''
        ws['F39'] = f"Cedula: {cedula}"

        # Embeber firma del usuario en F37
        second_signature = maintenance.second_signature
        if second_signature and second_signature.image:
            try:
                self._embed_image(ws, second_signature.image, 'F37', max_width=150, max_height=60)
            except Exception as e:
                print(f"No se pudo embeber firma usuario: {e}")

    def _embed_maintenance_photos(self, ws, maintenance):
        """Embebe fotos del mantenimiento al final del documento de forma ordenada."""
        photos = maintenance.photos.all()
        if not photos:
            return
        
        # Las fotos van al final, después de la fila 55, ordenadas verticalmente
        # Fila inicial para fotos: 56
        start_row = 56
        row_spacing = 12  # Espacio entre fotos (altura aproximada de cada foto)
        
        for idx, photo in enumerate(photos):
            if photo.image:
                try:
                    # Posicionar fotos verticalmente: A56, A68, A80, etc.
                    anchor_row = start_row + (idx * row_spacing)
                    anchor_cell = f'A{anchor_row}'
                    self._embed_image(ws, photo.image, anchor_cell, max_width=200, max_height=150)
                    
                    # Agregar descripción si existe
                    if photo.caption:
                        desc_row = anchor_row + 8  # Debajo de la foto
                        ws.cell(desc_row, 1).value = f"Foto {idx+1}: {photo.caption}"
                except Exception as e:
                    print(f"No se pudo embeber foto {idx+1}: {e}")

    def _embed_image(self, ws, image_field, anchor_cell: str, max_width: int = 150, max_height: int = 100):
        """
        Embebe una imagen en la hoja Excel.
        
        Args:
            ws: worksheet de openpyxl
            image_field: FileField/ImageField de Django con la imagen
            anchor_cell: celda donde anclar la imagen (ej: 'D40')
            max_width: ancho máximo en pixels
            max_height: alto máximo en pixels
        """
        # Leer imagen desde storage
        try:
            # Si es URL de storage, intentar abrir
            if hasattr(image_field, 'path'):
                # Local file
                img_path = image_field.path
                if os.path.exists(img_path):
                    pil_img = Image.open(img_path)
                else:
                    # Intentar leer desde storage
                    img_bytes = image_field.read()
                    pil_img = Image.open(BytesIO(img_bytes))
            else:
                # Desde storage
                img_bytes = image_field.read()
                pil_img = Image.open(BytesIO(img_bytes))
            
            # Redimensionar manteniendo aspect ratio
            pil_img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
            
            # Convertir a formato compatible con openpyxl
            img_buffer = BytesIO()
            pil_img.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            # Crear objeto imagen de openpyxl
            xl_img = XLImage(img_buffer)
            
            # Anclar en celda
            xl_img.anchor = anchor_cell
            
            # Agregar a worksheet
            ws.add_image(xl_img)
            
        except Exception as e:
            print(f"Error embebiendo imagen en {anchor_cell}: {e}")
            raise

    def _fill_service_rating(self, ws, maintenance):
        """Rellena la calificación del servicio."""
        calificacion = (maintenance.calificacion_servicio or '').lower().strip()
        
        # Las celdas de calificación están en fila 44/45: A=Excelente, C=Bueno, F=Regular, H=Malo
        # Marcar con ■ la calificación correspondiente
        # Buscar filas 44-45 para las marcas
        if calificacion in ('excelente', 'excellent', '5'):
            ws['A45'] = '■'  # O la celda correcta debajo de Excelente
        elif calificacion in ('bueno', 'good', '4'):
            ws['C45'] = '■'
        elif calificacion in ('regular', '3'):
            ws['F45'] = '■'
        elif calificacion in ('malo', 'bad', '1', '2'):
            ws['H45'] = '■'

    def _fill_user_observations(self, ws, maintenance):
        """Rellena las observaciones del usuario."""
        obs_usuario = maintenance.observaciones_usuario or ''
        # Fila 48 según el mapeo
        ws['A48'] = obs_usuario

        # Si hay observaciones adicionales, agregar en fila 52
        if maintenance.incident_notes:
            ws['A52'] = maintenance.incident_notes


def generate_excel_report(maintenance) -> bytes:
    """
    Función de conveniencia para generar un reporte Excel.
    
    Args:
        maintenance: Objeto Maintenance
        
    Returns:
        bytes: Contenido del archivo Excel
    """
    generator = ExcelReportGenerator()
    return generator.generate_report(maintenance)
