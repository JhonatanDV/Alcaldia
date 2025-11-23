import io
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
import openpyxl
from openpyxl.styles import Font, Alignment


class ExcelTemplateGenerator:
    def __init__(self, template_path: str | None = None):
        if template_path:
            self.workbook = openpyxl.load_workbook(template_path)
        else:
            self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active

    def fill_template(self, data: dict):
        """
        Rellena la plantilla con datos
        data = {
            'A1': 'Valor',
            'B2': 123,
            'cells': [
                {'cell': 'A5', 'value': 'Test', 'bold': True}
            ]
        }
        """
        for cell, value in data.items():
            if cell != 'cells':
                try:
                    self.sheet[cell] = value
                except Exception:
                    # ignore invalid cell addresses
                    pass

        if 'cells' in data and isinstance(data['cells'], list):
            for cell_data in data['cells']:
                try:
                    cell = self.sheet[cell_data['cell']]
                    cell.value = cell_data.get('value')

                    if cell_data.get('bold'):
                        cell.font = Font(bold=True)
                    if cell_data.get('align'):
                        cell.alignment = Alignment(horizontal=cell_data['align'])
                except Exception:
                    continue

    def create_table(self, start_row: int, headers: list, rows: list):
        for col_idx, header in enumerate(headers, start=1):
            cell = self.sheet.cell(row=start_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)

        for row_idx, row_data in enumerate(rows, start=start_row + 1):
            for col_idx, value in enumerate(row_data, start=1):
                self.sheet.cell(row=row_idx, column=col_idx, value=value)

    def get_file(self) -> io.BytesIO:
        virtual_file = io.BytesIO()
        self.workbook.save(virtual_file)
        virtual_file.seek(0)
        return virtual_file


class PDFTemplateGenerator:
    def __init__(self):
        self.buffer = io.BytesIO()
        self.pdf = canvas.Canvas(self.buffer, pagesize=letter)
        self.width, self.height = letter
        self.y_position = self.height - 50

    def add_header(self, title: str):
        self.pdf.setFont("Helvetica-Bold", 16)
        self.pdf.drawString(50, self.y_position, title)
        self.y_position -= 30

    def add_text(self, text: str, bold: bool = False):
        font = "Helvetica-Bold" if bold else "Helvetica"
        self.pdf.setFont(font, 12)
        self.pdf.drawString(50, self.y_position, text)
        self.y_position -= 20

    def add_table(self, headers: list, rows: list):
        x_start = 50
        col_width = 120

        self.pdf.setFont("Helvetica-Bold", 10)
        for i, header in enumerate(headers):
            self.pdf.drawString(x_start + (i * col_width), self.y_position, str(header))

        self.y_position -= 20

        self.pdf.setFont("Helvetica", 10)
        for row in rows:
            for i, cell in enumerate(row):
                try:
                    self.pdf.drawString(x_start + (i * col_width), self.y_position, str(cell))
                except Exception:
                    continue
            self.y_position -= 18
            if self.y_position < 50:
                self.pdf.showPage()
                self.y_position = self.height - 50

    def fill_template(self, data: dict):
        if 'title' in data:
            self.add_header(data['title'])

        if 'sections' in data and isinstance(data['sections'], list):
            for section in data['sections']:
                if section.get('type') == 'text':
                    self.add_text(section.get('content', ''), bold=section.get('bold', False))
                elif section.get('type') == 'table':
                    self.add_table(section.get('headers', []), section.get('rows', []))

    def get_file(self) -> io.BytesIO:
        self.pdf.save()
        self.buffer.seek(0)
        return self.buffer
