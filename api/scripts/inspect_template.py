import openpyxl
import sys

template_path = r'C:\Users\labinf1.pasto\Desktop\Alcaldia\plantillas\rutina_mantenimiento_preventivo_equipos_de_computo_.xlsx'

wb = openpyxl.load_workbook(template_path)
ws = wb.active

print('Hoja:', ws.title)
print('Dimensiones:', ws.dimensions)
print('\n=== Muestra de celdas clave ===')
for cell in ['A7', 'C7', 'H7', 'A8', 'C8', 'H8', 'A32', 'A34', 'A38', 'F38', 'A39', 'F39', 'A48']:
    val = ws[cell].value
    print(f'{cell}: {val}')

print('\n=== Celdas de actividades hardware (filas 11-14) ===')
for row in [11, 12, 13, 14]:
    a_val = ws.cell(row, 1).value
    d_val = ws.cell(row, 4).value
    e_val = ws.cell(row, 5).value
    f_val = ws.cell(row, 6).value
    i_val = ws.cell(row, 9).value
    j_val = ws.cell(row, 10).value
    print(f'Fila {row}: A="{a_val}" D="{d_val}" E="{e_val}" F="{f_val}" I="{i_val}" J="{j_val}"')

print('\n=== Celdas de actividades software (filas 18-22 muestra) ===')
for row in [18, 19, 20, 21, 22]:
    a_val = ws.cell(row, 1).value
    d_val = ws.cell(row, 4).value
    e_val = ws.cell(row, 5).value
    f_val = ws.cell(row, 6).value
    i_val = ws.cell(row, 9).value
    j_val = ws.cell(row, 10).value
    print(f'Fila {row}: A="{a_val}" D="{d_val}" E="{e_val}" F="{f_val}" I="{i_val}" J="{j_val}"')

print('\n=== Zonas de firma e imagen ===')
# Buscar celdas merged para firmas
print('Merged cells:', ws.merged_cells.ranges if hasattr(ws, 'merged_cells') else 'N/A')

# Buscar posibles ubicaciones de imagen (alrededor de filas 38-43 para firmas)
for row in range(38, 45):
    for col in [1, 6]:  # A y F
        val = ws.cell(row, col).value
        if val and ('nombre' in str(val).lower() or 'firma' in str(val).lower()):
            print(f'Celda {ws.cell(row, col).coordinate}: {val}')
