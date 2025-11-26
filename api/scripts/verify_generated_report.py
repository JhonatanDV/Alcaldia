import openpyxl
import sys

# Abrir el archivo generado
file_path = r'C:\Users\labinf1.pasto\Desktop\Alcaldia\media\maintenance-reports\rutina_mantenimiento_EQ-0003_20251125_2acd397e.xlsx'

wb = openpyxl.load_workbook(file_path)
ws = wb.active

print(f"=== VERIFICACIÓN DEL REPORTE GENERADO ===\n")
print(f"Archivo: {file_path}")
print(f"Hoja: {ws.title}\n")

print("=== ENCABEZADO ===")
print(f"A7 (Sede): {ws['A7'].value}")
print(f"C7 (Dependencia): {ws['C7'].value}")
print(f"H7 (Oficina): {ws['H7'].value}")
print(f"A8 (Placa): {ws['A8'].value}")
print(f"C8 (Fecha): {ws['C8'].value}")
print(f"H8 (Horas): {ws['H8'].value}")

print("\n=== ACTIVIDADES HARDWARE (muestra filas 11-14) ===")
for row in [11, 12, 13, 14]:
    a_val = ws.cell(row, 1).value
    d_val = ws.cell(row, 4).value or ''
    e_val = ws.cell(row, 5).value or ''
    f_val = ws.cell(row, 6).value
    i_val = ws.cell(row, 9).value or ''
    j_val = ws.cell(row, 10).value or ''
    print(f"Fila {row}:")
    print(f"  Izq: {a_val[:50] if a_val else ''} -> SI:{d_val} NA:{e_val}")
    print(f"  Der: {f_val[:50] if f_val else ''} -> SI:{i_val} NA:{j_val}")

print("\n=== ACTIVIDADES SOFTWARE (muestra filas 18-22) ===")
for row in [18, 19, 20, 21, 22]:
    a_val = ws.cell(row, 1).value
    d_val = ws.cell(row, 4).value or ''
    e_val = ws.cell(row, 5).value or ''
    f_val = ws.cell(row, 6).value
    i_val = ws.cell(row, 9).value or ''
    j_val = ws.cell(row, 10).value or ''
    print(f"Fila {row}:")
    print(f"  Izq: {a_val[:50] if a_val else ''} -> SI:{d_val} NA:{e_val}")
    if f_val:
        print(f"  Der: {f_val[:50]} -> SI:{i_val} NA:{j_val}")

print("\n=== OBSERVACIONES ===")
print(f"A32 (Obs Generales): {ws['A32'].value[:100] if ws['A32'].value else 'Vacío'}...")
print(f"A34 (Obs Seguridad): {ws['A34'].value[:100] if ws['A34'].value else 'Vacío'}...")

print("\n=== FIRMAS ===")
print(f"A38 (Nombre Técnico): {ws['A38'].value}")
print(f"A39 (Cargo): {ws['A39'].value}")
print(f"F38 (Nombre Usuario): {ws['F38'].value}")
print(f"F39 (Cédula): {ws['F39'].value}")

print("\n=== CALIFICACIÓN SERVICIO (fila 45) ===")
print(f"A45 (Excelente): '{ws['A45'].value}'")
print(f"C45 (Bueno): '{ws['C45'].value}'")
print(f"F45 (Regular): '{ws['F45'].value}'")
print(f"H45 (Malo): '{ws['H45'].value}'")

print("\n=== IMÁGENES EMBEBIDAS ===")
print(f"Número de imágenes en el worksheet: {len(ws._images)}")
for idx, img in enumerate(ws._images):
    anchor_cell = img.anchor._from.col if hasattr(img.anchor, '_from') else 'N/A'
    print(f"  Imagen {idx+1}: tipo={type(img).__name__}")

print("\n✅ Verificación completada")
