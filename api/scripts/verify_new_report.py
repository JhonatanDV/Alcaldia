import openpyxl
import sys

# Abrir el archivo generado
file_path = r'C:\Users\labinf1.pasto\Desktop\Alcaldia\media\maintenance-reports\rutina_mantenimiento_EQ-0003_20251125_c1852b13.xlsx'

wb = openpyxl.load_workbook(file_path)
ws = wb.active

print(f"=== VERIFICACIÓN DEL NUEVO REPORTE ===\n")
print(f"Archivo: {file_path}\n")

print("=== ENCABEZADO ===")
print(f"A7 (Sede): {ws['A7'].value}")
print(f"C7 (Dependencia): {ws['C7'].value}")
print(f"H7 (Oficina): {ws['H7'].value}")
print(f"C8 (Fecha): {ws['C8'].value}")
print(f"H8 (Horas): {ws['H8'].value}")

print("\n=== ZONA DE FIRMAS (filas 37-39) ===")
print(f"A37 (Firma técnico): {ws['A37'].value}")
print(f"F37 (Firma usuario): {ws['F37'].value}")
print(f"A38 (Nombre técnico): {ws['A38'].value}")
print(f"F38 (Nombre usuario): {ws['F38'].value}")
print(f"A39 (Cargo técnico): {ws['A39'].value}")
print(f"F39 (Cédula usuario): {ws['F39'].value}")

print("\n=== IMÁGENES EMBEBIDAS ===")
print(f"Total de imágenes: {len(ws._images)}")

# Analizar posiciones de imágenes
firma_images = []
photo_images = []

for idx, img in enumerate(ws._images):
    # Obtener posición aproximada del anchor
    try:
        if hasattr(img, 'anchor'):
            anchor_str = str(img.anchor)
            # Intentar extraer celda de inicio
            if hasattr(img.anchor, '_from'):
                from_obj = img.anchor._from
                if hasattr(from_obj, 'row'):
                    row_num = from_obj.row
                    col_num = from_obj.col
                    
                    # Clasificar según posición
                    if row_num <= 40:
                        firma_images.append((idx+1, row_num, col_num))
                        print(f"  Imagen {idx+1}: FIRMA (fila ~{row_num}, col ~{col_num})")
                    else:
                        photo_images.append((idx+1, row_num, col_num))
                        print(f"  Imagen {idx+1}: FOTO (fila ~{row_num}, col ~{col_num})")
                else:
                    print(f"  Imagen {idx+1}: tipo={type(img).__name__}")
            else:
                print(f"  Imagen {idx+1}: anchor={type(img.anchor).__name__}")
    except Exception as e:
        print(f"  Imagen {idx+1}: No se pudo determinar posición ({e})")

print(f"\n📍 Resumen:")
print(f"   - Firmas (fila ≤40): {len(firma_images)}")
print(f"   - Fotos (fila >40): {len(photo_images)}")

print("\n=== ZONA FINAL (filas 50-70 para fotos) ===")
for row in range(50, 71):
    val = ws.cell(row, 1).value
    if val and len(str(val)) > 2:
        print(f"Fila {row}: {str(val)[:60]}")

print("\n✅ Verificación completada")
