import openpyxl
import sys

template_path = r'C:\Users\labinf1.pasto\Desktop\Alcaldia\plantillas\rutina_mantenimiento_preventivo_equipos_de_computo_.xlsx'
mapeo_path = r'C:\Users\labinf1.pasto\Desktop\Alcaldia\plantillas\Mapeo_texto_rutina_mantenimiento_preventivo_equipos_de_computo_.xlsx'

print("=== PLANTILLA ORIGINAL ===")
wb_template = openpyxl.load_workbook(template_path)
ws_template = wb_template.active

print("\n=== ARCHIVO DE MAPEO ===")
wb_mapeo = openpyxl.load_workbook(mapeo_path)
ws_mapeo = wb_mapeo.active

# Comparar celdas para encontrar diferencias (zonas marcadas)
print("\n=== DIFERENCIAS (Zonas marcadas para rellenar) ===")
differences = []

for row in range(1, min(60, ws_template.max_row + 1)):
    for col in range(1, 11):  # A-J
        cell_template = ws_template.cell(row, col)
        cell_mapeo = ws_mapeo.cell(row, col)
        
        val_template = cell_template.value
        val_mapeo = cell_mapeo.value
        
        if val_template != val_mapeo:
            cell_coord = ws_template.cell(row, col).coordinate
            differences.append({
                'cell': cell_coord,
                'row': row,
                'col': col,
                'original': val_template,
                'mapeo': val_mapeo
            })

print(f"\nTotal de diferencias encontradas: {len(differences)}")
print("\n=== CELDAS A RELLENAR (con marcadores en mapeo) ===")

# Agrupar por secciones
header_cells = [d for d in differences if d['row'] <= 10]
hardware_cells = [d for d in differences if 11 <= d['row'] <= 16]
software_cells = [d for d in differences if 17 <= d['row'] <= 31]
obs_cells = [d for d in differences if 32 <= d['row'] <= 36]
firma_cells = [d for d in differences if 37 <= d['row'] <= 43]
rating_cells = [d for d in differences if 44 <= d['row'] <= 47]
final_cells = [d for d in differences if d['row'] >= 48]

print(f"\nENCABEZADO (filas 1-10): {len(header_cells)} celdas")
for d in header_cells[:10]:
    print(f"  {d['cell']}: '{d['original']}' -> '{d['mapeo']}'")

print(f"\nHARDWARE (filas 11-16): {len(hardware_cells)} celdas")
for d in hardware_cells[:5]:
    print(f"  {d['cell']}: marcador='{d['mapeo']}'")

print(f"\nSOFTWARE (filas 17-31): {len(software_cells)} celdas")
for d in software_cells[:5]:
    print(f"  {d['cell']}: marcador='{d['mapeo']}'")

print(f"\nOBSERVACIONES (filas 32-36): {len(obs_cells)} celdas")
for d in obs_cells:
    print(f"  {d['cell']}: marcador='{d['mapeo']}'")

print(f"\nFIRMAS (filas 37-43): {len(firma_cells)} celdas")
for d in firma_cells:
    print(f"  {d['cell']}: '{d['original']}' -> '{d['mapeo']}'")

print(f"\nCALIFICACIÓN (filas 44-47): {len(rating_cells)} celdas")
for d in rating_cells[:5]:
    print(f"  {d['cell']}: marcador='{d['mapeo']}'")

print(f"\nFINAL/FOTOS (filas 48+): {len(final_cells)} celdas")
for d in final_cells[:5]:
    print(f"  {d['cell']}: marcador='{d['mapeo']}'")

# Verificar zona de firmas específicamente
print("\n=== VERIFICACIÓN ZONA DE FIRMAS ===")
print(f"A37: Template='{ws_template['A37'].value}' | Mapeo='{ws_mapeo['A37'].value}'")
print(f"F37: Template='{ws_template['F37'].value}' | Mapeo='{ws_mapeo['F37'].value}'")
print(f"A38: Template='{ws_template['A38'].value}' | Mapeo='{ws_mapeo['A38'].value}'")
print(f"F38: Template='{ws_template['F38'].value}' | Mapeo='{ws_mapeo['F38'].value}'")
print(f"A39: Template='{ws_template['A39'].value}' | Mapeo='{ws_mapeo['A39'].value}'")
print(f"F39: Template='{ws_template['F39'].value}' | Mapeo='{ws_mapeo['F39'].value}'")

# Buscar zona final para fotos
print("\n=== ZONA FINAL PARA FOTOS (filas 48-60) ===")
for row in range(48, 61):
    a_val = ws_template.cell(row, 1).value
    if a_val:
        print(f"Fila {row}: {a_val}")
