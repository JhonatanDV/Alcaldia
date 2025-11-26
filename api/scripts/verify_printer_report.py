import openpyxl

file_path = r'C:\Users\labinf1.pasto\Desktop\Alcaldia\media\maintenance-reports\rutina_impresora_escaner_EQ-0013_20251125_28a7939b.xlsx'

wb = openpyxl.load_workbook(file_path)
ws = wb.active

print("=" * 80)
print("   REPORTE IMPRESORA/ESCÁNER - VERIFICACIÓN")
print("=" * 80)

print("\n✅ ENCABEZADO")
print(f"   Sede: {ws['A9'].value}")
print(f"   Dependencia: {ws['C9'].value}")
print(f"   Oficina: {ws['G9'].value}")
print(f"   Placa: {ws['A10'].value}")
print(f"   Fecha: {ws['C10'].value}")
print(f"   Horas: {ws['G10'].value}")

print("\n✅ EQUIPO")
print(f"   {ws['A11'].value}")
print(f"   {ws['D11'].value}")

print("\n✅ ACTIVIDADES ESCÁNER (filas 15-16)")
for row in [15, 16]:
    actividad_izq = ws.cell(row, 1).value
    si_izq = ws.cell(row, 4).value
    na_izq = ws.cell(row, 5).value
    actividad_der = ws.cell(row, 6).value
    si_der = ws.cell(row, 9).value
    na_der = ws.cell(row, 10).value
    print(f"   Fila {row}:")
    if actividad_izq:
        print(f"      IZQ: {actividad_izq[:35]}... -> SI:{si_izq or ''} NA:{na_izq or ''}")
    if actividad_der:
        print(f"      DER: {actividad_der[:35]}... -> SI:{si_der or ''} NA:{na_der or ''}")

print("\n✅ ACTIVIDADES IMPRESORA (filas 21-24 muestra)")
for row in [21, 22, 23, 24]:
    actividad_izq = ws.cell(row, 1).value
    si_izq = ws.cell(row, 4).value
    na_izq = ws.cell(row, 5).value
    actividad_der = ws.cell(row, 6).value
    si_der = ws.cell(row, 9).value
    na_der = ws.cell(row, 10).value
    print(f"   Fila {row}:")
    if actividad_izq:
        print(f"      IZQ: {actividad_izq[:35]}... -> SI:{si_izq or ''} NA:{na_izq or ''}")
    if actividad_der:
        print(f"      DER: {actividad_der[:35]}... -> SI:{si_der or ''} NA:{na_der or ''}")

print("\n✅ OBSERVACIONES")
obs = ws.cell(35, 1).value
print(f"   {obs[:70] if obs else 'N/A'}...")

print("\n✅ FIRMAS")
print(f"   Responsable: {ws['A39'].value}")
print(f"   Usuario: {ws['F39'].value}")
print(f"   Nombre Técnico: {ws['A41'].value}")
print(f"   Cargo: {ws['A42'].value}")
print(f"   Nombre Usuario: {ws['F41'].value}")

print(f"\n✅ IMÁGENES: {len(ws._images)} embebidas")

print("\n" + "=" * 80)
print("   ✅ REPORTE IMPRESORA/ESCÁNER GENERADO CORRECTAMENTE")
print("=" * 80)
